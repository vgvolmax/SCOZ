import re
import zipfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from backend.domain.product_snapshot import (
    CategoryMismatch, ConflictingObservationRows, IncompatibleReportSchema,
    InvalidMetricValue, InvalidProductIdentity, InvalidReportPeriod,
    ParsedOzonProductRow, ParsedOzonProductsReport, RowError, UnsupportedWorkbook,
    decimal_from_excel_number,
    WrongReportType, snapshot_payload_sha256,
)

HEADERS = ("Название товара","Ссылка на товар","Продавец","Бренд","Категория 1 уровня","Категория 3 уровня","Признак товара","Заказано на сумму, ₽","Динамика оборота, %","Заказано, штуки","Средняя цена, ₽","Минимальная цена, ₽","Доля выкупа, %","Упущенные продажи","Дней без остатка","Среднесуточные продажи, ₽","Среднесуточные продажи, штуки","Остаток на конец периода, штуки","Схема работы","Объем товара, л","Показы всего","Просмотры в поиске и каталоге","Просмотры карточки","Конверсия из показа в заказ, %","В корзину из поиска и каталога, %","В корзину из карточки, %","Скидка за счет акций","Доля суммы заказов по акциям, %","Дней в акциях","Дней с продвижением","Доля рекламных расходов, %","Дата создания карточки товара")
URL_RE = re.compile(r"https://www\.ozon\.ru/product/(\d+)/?")
WINDOW_RE = re.compile(r"([1-9][0-9]*) дней")
DAY_RE = re.compile(r"([0-9]+) из ([1-9][0-9]*)")

def _formula(value: object) -> bool: return isinstance(value, str) and value.startswith("=")
def _text(value: object) -> str:
    if _formula(value) or not isinstance(value, str) or not value: raise InvalidMetricValue()
    return value
def _decimal(value: object, *, sentinel: bool = False) -> Decimal | None:
    if sentinel and value == "Нет данных": return None
    return decimal_from_excel_number(value)
def _count(value: object) -> int:
    number = _decimal(value)
    assert number is not None
    if number < 0 or number != number.to_integral_value(): raise InvalidMetricValue()
    return int(number)
def _days(value: object, *, optional: bool = False) -> tuple[int|None,int|None]:
    if optional and value == "-": return None,None
    if not isinstance(value,str) or not (match := DAY_RE.fullmatch(value)): raise InvalidMetricValue()
    numerator,denominator = map(int,match.groups())
    if numerator > denominator: raise InvalidMetricValue()
    return numerator,denominator
def _date(value: object) -> date:
    if not isinstance(value,str): raise InvalidMetricValue()
    try: return date.fromisoformat(value)
    except ValueError as error: raise InvalidMetricValue() from error

def parse_ozon_products_xlsx(path: Path) -> ParsedOzonProductsReport:
    try:
        with zipfile.ZipFile(path) as package:
            worksheet_xml = [name for name in package.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")]
            if any(b"<mergeCells" in package.read(name) for name in worksheet_xml):
                raise IncompatibleReportSchema()
    except IncompatibleReportSchema:
        raise
    except (OSError, zipfile.BadZipFile, KeyError) as error:
        raise UnsupportedWorkbook() from error
    try:
        source = path.open("rb")
        workbook = load_workbook(filename=source,read_only=True,data_only=False)
    except Exception as error:
        if "source" in locals():
            source.close()
        raise UnsupportedWorkbook() from error
    try:
        if len(workbook.worksheets) != 1: raise IncompatibleReportSchema()
        sheet = workbook.worksheets[0]
        # Current Ozon packages may advertise only A1 although cells through AF
        # are physically present.  Re-scan the stream; exact headers below remain
        # the authoritative business-schema check.
        sheet.reset_dimensions()
        sheet.calculate_dimension(force=True)
        if sheet.max_column != 32: raise IncompatibleReportSchema()
        iterator = sheet.iter_rows(min_row=1, max_col=32)
        prefix = []
        for _ in range(6):
            try:
                prefix.append(next(iterator))
            except StopIteration:
                raise IncompatibleReportSchema()
        markers = tuple(prefix[i][0].value for i in range(3))
        expected = ("Дата формирования:","Период отчета:","Категория 3 уровня:")
        if markers != expected:
            if markers[0] != expected[0]: raise WrongReportType()
            raise IncompatibleReportSchema()
        structure = [cell.value for row in prefix for cell in row]
        if any(_formula(v) for v in structure): raise IncompatibleReportSchema()
        if any(cell.value is not None for cell in prefix[3]): raise IncompatibleReportSchema()
        if tuple(cell.value for cell in prefix[4]) != HEADERS or prefix[5][0].value != "Среднее значение по товарам": raise IncompatibleReportSchema()
        try:
            generated = datetime.strptime(prefix[0][1].value,"%m.%d.%y").date()
            match = WINDOW_RE.fullmatch(prefix[1][1].value)
            if not match: raise ValueError
            window = int(match.group(1))
        except (TypeError,ValueError): raise InvalidReportPeriod()
        report_category = prefix[2][1].value
        if not isinstance(report_category, str) or not report_category:
            raise IncompatibleReportSchema()
        rows=[]; errors=[]; seen=duplicates=warnings=0; unique={}
        for row_number, row in enumerate(iterator, start=7):
            cells=[cell.value for cell in row]
            if all(value is None for value in cells): continue
            seen += 1
            try:
                url = cells[1]
                if not isinstance(url,str) or not (match := URL_RE.fullmatch(url)): raise InvalidProductIdentity()
                identity=match.group(1)
                if cells[5] != report_category: raise CategoryMismatch()
                badge = None if cells[6] in (None,"") else _text(cells[6])
                out_days,out_window=_days(cells[14],optional=True)
                promotion_days,promotion_window=_days(cells[28]); advertising_days,advertising_window=_days(cells[29])
                values=dict(product_url=f"https://www.ozon.ru/product/{identity}",title=_text(cells[0]),seller_name=_text(cells[2]),brand=_text(cells[3]),category_level_1=_text(cells[4]),category_level_3=_text(cells[5]),product_badges=badge,ordered_amount_rub=_decimal(cells[7]),turnover_change_pct=_decimal(cells[8],sentinel=True),ordered_units=_count(cells[9]),average_price_rub=_decimal(cells[10]),minimum_price_rub=_decimal(cells[11]),buyout_share_pct=_decimal(cells[12],sentinel=True),missed_sales_source_value=_decimal(cells[13]),out_of_stock_days=out_days,out_of_stock_window_days=out_window,avg_daily_sales_rub=_decimal(cells[15]),avg_daily_sales_units=_count(cells[16]),stock_end_units=_count(cells[17]),fulfillment_scheme=_text(cells[18]),volume_l=_decimal(cells[19]),impressions_total=_count(cells[20]),search_catalog_views=_count(cells[21]),card_views=_count(cells[22]),impression_to_order_pct=_decimal(cells[23]),search_catalog_to_cart_pct=_decimal(cells[24]),card_to_cart_pct=_decimal(cells[25]),promotion_discount_source_value=_decimal(cells[26]),promotion_order_amount_share_pct=_decimal(cells[27]),promotion_days=promotion_days,promotion_window_days=promotion_window,advertising_days=advertising_days,advertising_window_days=advertising_window,total_drr_pct=_decimal(cells[30]),card_created_on=_date(cells[31]))
                payload_hash=snapshot_payload_sha256(values); key=(identity,generated,window)
                if key in unique:
                    if unique[key] != payload_hash: raise ConflictingObservationRows()
                    duplicates += 1; warnings += 1; continue
                unique[key]=payload_hash; rows.append(ParsedOzonProductRow(row_number,identity,values,payload_hash))
            except ConflictingObservationRows: raise
            except (InvalidProductIdentity,CategoryMismatch,InvalidMetricValue) as error:
                messages = {
                    InvalidProductIdentity: "Некорректная ссылка на товар.",
                    InvalidMetricValue: "Некорректное значение показателя.",
                    CategoryMismatch: "Категория товара не совпадает с категорией отчёта.",
                }
                errors.append(RowError(row_number,type(error).__name__,messages[type(error)]))
        if not rows: raise InvalidMetricValue("report contains no usable rows")
        return ParsedOzonProductsReport(generated,window,seen,tuple(rows),tuple(errors),duplicates,warnings)
    finally:
        workbook.close()
        source.close()
