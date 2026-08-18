"""Strict parser for the verified Ozon ``explainer_report`` XLSX shape."""

from __future__ import annotations

import re
import zipfile
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from backend.domain.search_visibility import (
    CpoState,
    ParsedSearchVisibilityReport,
    ParsedSearchVisibilityRow,
    SearchVisibilityConflictingObservationRows,
    SearchVisibilityIncompatibleReportSchema,
    SearchVisibilityInvalidMetricValue,
    SearchVisibilityInvalidObservedAt,
    SearchVisibilityInvalidProductIdentity,
    SearchVisibilityInvalidSearchContext,
    SearchVisibilityRowError,
    SearchVisibilityUnsupportedWorkbook,
    SearchVisibilityWrongReportType,
    search_visibility_payload_sha256,
)


HEADERS = (
    "Позиция",
    "ID товара",
    "Название товара",
    "Имя селлера",
    "Сводная оценка",
    "Статус",
    "Ставка\nОплата за клик",
    "Стратегия",
    "Ставка\nОплата за заказ",
    "Соответствие запросу",
    "Отзывы",
    "Цена для покупателя",
    "Популярность общая",
    "Акции от Ozon",
    "Срок доставки",
    "Индекс цен",
)

_DATE_RE = re.compile(r"Дата: ([0-9]{2}/[0-9]{2}/[0-9]{4})")
_TIME_RE = re.compile(r"Время: ([0-9]{2}:[0-9]{2}) \+00")
_DECLARED_RE = re.compile(r"Сколько позиций в выдаче: ([1-9][0-9]*)")
_POSITION_RE = re.compile(r"[1-9][0-9]*")
_DECIMAL_RE = re.compile(r"[0-9]+,[0-9]+")
_MONEY_2_RE = re.compile(r"((?:[0-9]+|[1-9][0-9]{0,2}(?: [0-9]{3})+),[0-9]{2}) ₽")
_WHOLE_MONEY_RE = re.compile(r"((?:[0-9]+|[1-9][0-9]{0,2}(?: [0-9]{3})+)) ₽")
_CPO_RE = re.compile(r"([0-9]+)%")
_REVIEWS_RE = re.compile(r"([0-9]+,[0-9]+) \(((?:[0-9]+|[1-9][0-9]{0,2}(?: [0-9]{3})+)) шт\.\)")
_DELIVERY_RE = re.compile(r"([0-9]+)-([0-9]+) (день|дня|дней)")
_PRICE_INDEX_RE = re.compile(r"([0-9]+,[0-9]+)%")
_PERIOD_START_RE = re.compile(r"Дата начала: [0-9]{2}/[0-9]{2}/[0-9]{4}")
_PERIOD_END_RE = re.compile(r"Дата конца: [0-9]{2}/[0-9]{2}/[0-9]{4}")
_SELLER_QUERIES_HEADERS = ("SKU", "Артикул", "Название товара", "Запросы товара")


class _RowProblem(ValueError):
    def __init__(self, code: str, message: str) -> None:
        self.code = code
        self.message = message
        super().__init__(message)


def _edge_cleanup_identity(value: object) -> str:
    if not isinstance(value, str):
        raise SearchVisibilityInvalidSearchContext()
    cleaned = value.strip(" \u00a0")
    if not cleaned:
        raise SearchVisibilityInvalidSearchContext()
    return cleaned


def _parse_observed_at(date_value: object, time_value: object) -> datetime:
    if not isinstance(date_value, str) or not isinstance(time_value, str):
        raise SearchVisibilityInvalidObservedAt()
    date_match = _DATE_RE.fullmatch(date_value)
    time_match = _TIME_RE.fullmatch(time_value)
    if not date_match or not time_match:
        raise SearchVisibilityInvalidObservedAt()
    try:
        return datetime.strptime(
            f"{date_match.group(1)} {time_match.group(1)}", "%d/%m/%Y %H:%M"
        ).replace(tzinfo=timezone.utc)
    except ValueError as error:
        raise SearchVisibilityInvalidObservedAt() from error


def _semantically_blank(value: object) -> bool:
    return value is None or value == ""


def _formula_cell(cell: object) -> bool:
    return getattr(cell, "data_type", None) == "f" or (
        isinstance(getattr(cell, "value", None), str)
        and getattr(cell, "value").startswith("=")
    )


def _looks_like_seller_queries(sheet: object) -> bool:
    """Recognize the known seller-queries report without relying on its name."""
    metadata = tuple(sheet.cell(row=row, column=1).value for row in range(1, 5))
    return (
        isinstance(metadata[0], str)
        and _DATE_RE.fullmatch(metadata[0]) is not None
        and isinstance(metadata[1], str)
        and _TIME_RE.fullmatch(metadata[1]) is not None
        and isinstance(metadata[2], str)
        and _PERIOD_START_RE.fullmatch(metadata[2]) is not None
        and isinstance(metadata[3], str)
        and _PERIOD_END_RE.fullmatch(metadata[3]) is not None
        and all(
            _semantically_blank(sheet.cell(row=5, column=column).value)
            for column in range(1, 17)
        )
        and tuple(
            sheet.cell(row=6, column=column).value
            for column in range(1, len(_SELLER_QUERIES_HEADERS) + 1)
        )
        == _SELLER_QUERIES_HEADERS
    )


def _required_text(value: object) -> str:
    if not isinstance(value, str) or value == "":
        raise _RowProblem("INVALID_METRIC_VALUE", "Некорректное значение показателя.")
    return value


def _parse_position(value: object) -> int:
    if not isinstance(value, str) or not _POSITION_RE.fullmatch(value):
        raise _RowProblem("INVALID_POSITION", "Некорректная позиция товара.")
    return int(value)


def _parse_product_id(value: object) -> str:
    if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
        raise _RowProblem("INVALID_PRODUCT_IDENTITY", "Некорректный ID товара.")
    return str(value)


def _decimal_text(source: str) -> Decimal:
    try:
        return Decimal(source.replace(" ", "").replace(",", "."))
    except InvalidOperation as error:  # defensive; regular expressions own the grammar
        raise SearchVisibilityInvalidMetricValue() from error


def _parse_decimal_comma(value: object) -> Decimal:
    if not isinstance(value, str) or not _DECIMAL_RE.fullmatch(value):
        raise _RowProblem("INVALID_METRIC_VALUE", "Некорректное значение показателя.")
    return _decimal_text(value)


def _parse_cpc(value: object) -> Decimal:
    match = _MONEY_2_RE.fullmatch(value) if isinstance(value, str) else None
    if not match:
        raise _RowProblem("INVALID_METRIC_VALUE", "Некорректное значение показателя.")
    return _decimal_text(match.group(1))


def _parse_cpo(value: object) -> tuple[CpoState, Decimal | None]:
    if value == "Выключено":
        return CpoState.DISABLED, None
    if value == "—":
        return CpoState.UNAVAILABLE, None
    match = _CPO_RE.fullmatch(value) if isinstance(value, str) else None
    if not match:
        raise _RowProblem("INVALID_CPO_STATE", "Некорректное значение оплаты за заказ.")
    return CpoState.ACTIVE, Decimal(match.group(1))


def _parse_reviews(value: object) -> tuple[Decimal | None, int | None]:
    if value == "— ":
        return None, None
    match = _REVIEWS_RE.fullmatch(value) if isinstance(value, str) else None
    if not match:
        raise _RowProblem("INVALID_REVIEWS", "Некорректное значение отзывов.")
    return _decimal_text(match.group(1)), int(match.group(2).replace(" ", ""))


def _parse_buyer_price(value: object) -> Decimal:
    match = _WHOLE_MONEY_RE.fullmatch(value) if isinstance(value, str) else None
    if not match:
        raise _RowProblem("INVALID_METRIC_VALUE", "Некорректное значение показателя.")
    return _decimal_text(match.group(1))


def _parse_delivery(value: object) -> tuple[str, int, int]:
    match = _DELIVERY_RE.fullmatch(value) if isinstance(value, str) else None
    if not match or int(match.group(1)) > int(match.group(2)):
        raise _RowProblem("INVALID_DELIVERY", "Некорректный срок доставки.")
    return value, int(match.group(1)), int(match.group(2))


def _parse_price_index(value: object) -> Decimal:
    match = _PRICE_INDEX_RE.fullmatch(value) if isinstance(value, str) else None
    if not match:
        raise _RowProblem("INVALID_METRIC_VALUE", "Некорректное значение показателя.")
    return _decimal_text(match.group(1))


def _parse_product_row(cells: list[object]) -> tuple[str, dict[str, object]]:
    product_id = _parse_product_id(cells[1])
    cpo_state, cpo_pct = _parse_cpo(cells[8])
    rating, reviews_count = _parse_reviews(cells[10])
    delivery_label, delivery_min, delivery_max = _parse_delivery(cells[14])
    if cells[13] == "Да":
        ozon_promotion = True
    elif cells[13] == "Нет":
        ozon_promotion = False
    else:
        raise _RowProblem("INVALID_METRIC_VALUE", "Некорректное значение показателя.")
    values: dict[str, object] = {
        "source_title": _required_text(cells[2]),
        "seller_name": _required_text(cells[3]),
        "position": _parse_position(cells[0]),
        "overall_score": _parse_decimal_comma(cells[4]),
        "promotion_status": _required_text(cells[5]),
        "cpc_rub": _parse_cpc(cells[6]),
        "promotion_strategy": _required_text(cells[7]),
        "cpo_state": cpo_state,
        "cpo_pct": cpo_pct,
        "relevance_score": _parse_decimal_comma(cells[9]),
        "rating": rating,
        "reviews_count": reviews_count,
        "buyer_price_rub": _parse_buyer_price(cells[11]),
        "popularity_score": _parse_decimal_comma(cells[12]),
        "ozon_promotion": ozon_promotion,
        "delivery_label": delivery_label,
        "delivery_min_days": delivery_min,
        "delivery_max_days": delivery_max,
        "price_index_pct": _parse_price_index(cells[15]),
    }
    return product_id, values


def _has_merged_cells(path: Path) -> bool:
    try:
        with zipfile.ZipFile(path) as package:
            return any(
                b"<mergeCells" in package.read(name)
                for name in package.namelist()
                if name.startswith("xl/worksheets/") and name.endswith(".xml")
            )
    except (OSError, zipfile.BadZipFile, KeyError) as error:
        raise SearchVisibilityUnsupportedWorkbook() from error


def parse_ozon_search_visibility_xlsx(path: Path) -> ParsedSearchVisibilityReport:
    """Parse only Source Contract v1, without filename or sheet-name guessing."""
    if _has_merged_cells(path):
        raise SearchVisibilityIncompatibleReportSchema()
    source = None
    workbook = None
    try:
        source = path.open("rb")
        workbook = load_workbook(filename=source, read_only=True, data_only=False)
    except Exception as error:
        if source is not None:
            source.close()
        raise SearchVisibilityUnsupportedWorkbook() from error

    try:
        if len(workbook.worksheets) != 1:
            raise SearchVisibilityIncompatibleReportSchema()
        sheet = workbook.worksheets[0]
        structural = [sheet.cell(row=row, column=column) for row in range(1, 10) for column in range(1, 17)]
        if any(_formula_cell(cell) for cell in structural):
            raise SearchVisibilityIncompatibleReportSchema()

        metadata = tuple(sheet.cell(row=row, column=1).value for row in range(1, 6))
        expected_prefixes = ("Дата: ", "Запрос: ", "Время: ", "Регион: ", "Сколько позиций в выдаче: ")
        markers = tuple(isinstance(value, str) and value.startswith(prefix) for value, prefix in zip(metadata, expected_prefixes, strict=True))
        if _looks_like_seller_queries(sheet):
            raise SearchVisibilityWrongReportType()
        if not any(markers):
            raise SearchVisibilityWrongReportType()
        if not all(markers):
            raise SearchVisibilityIncompatibleReportSchema()
        if any(not _semantically_blank(sheet.cell(row=row, column=column).value) for row in (6, 8) for column in range(1, 17)):
            raise SearchVisibilityIncompatibleReportSchema()
        if tuple(sheet.cell(row=7, column=column).value for column in range(1, 17)) != HEADERS:
            raise SearchVisibilityIncompatibleReportSchema()
        # Physical formatting through Z is allowed; business values past P are not.
        if any(
            not _semantically_blank(sheet.cell(row=row, column=column).value)
            for row in range(1, sheet.max_row + 1)
            for column in range(17, sheet.max_column + 1)
        ):
            raise SearchVisibilityIncompatibleReportSchema()

        observed_at = _parse_observed_at(metadata[0], metadata[2])
        query_text = _edge_cleanup_identity(metadata[1][len(expected_prefixes[1]) :])
        cluster_name = _edge_cleanup_identity(metadata[3][len(expected_prefixes[3]) :])
        declared_match = _DECLARED_RE.fullmatch(metadata[4])
        if not declared_match:
            raise SearchVisibilityIncompatibleReportSchema()
        declared_rows = int(declared_match.group(1))

        candidates: list[tuple[int, list[object], list[object]]] = []
        for row_number in range(10, sheet.max_row + 1):
            cell_objects = [sheet.cell(row=row_number, column=column) for column in range(1, 17)]
            values = [cell.value for cell in cell_objects]
            if all(_semantically_blank(value) for value in values):
                continue
            candidates.append((row_number, values, cell_objects))
        if len(candidates) != declared_rows:
            raise SearchVisibilityIncompatibleReportSchema()

        rows: list[ParsedSearchVisibilityRow] = []
        row_errors: list[SearchVisibilityRowError] = []
        unique: dict[str, str] = {}
        duplicate_input_rows = 0
        warnings_count = 0
        for row_number, values, cell_objects in candidates:
            try:
                if any(_formula_cell(cell) for cell in cell_objects):
                    raise _RowProblem("INVALID_METRIC_VALUE", "Некорректное значение показателя.")
                product_id, snapshot_values = _parse_product_row(values)
                payload_sha256 = search_visibility_payload_sha256(snapshot_values)
                prior_hash = unique.get(product_id)
                if prior_hash is not None:
                    if prior_hash != payload_sha256:
                        raise SearchVisibilityConflictingObservationRows()
                    duplicate_input_rows += 1
                    warnings_count += 1
                    continue
                unique[product_id] = payload_sha256
                rows.append(ParsedSearchVisibilityRow(row_number, product_id, snapshot_values, payload_sha256))
            except SearchVisibilityConflictingObservationRows:
                raise
            except _RowProblem as error:
                row_errors.append(SearchVisibilityRowError(row_number, error.code, error.message))

        return ParsedSearchVisibilityReport(
            observed_at,
            query_text,
            cluster_name,
            declared_rows,
            len(candidates),
            tuple(rows),
            tuple(row_errors),
            duplicate_input_rows,
            warnings_count,
        )
    finally:
        workbook.close()
        source.close()


__all__ = ["parse_ozon_search_visibility_xlsx"]
