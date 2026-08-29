from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path, PurePosixPath
import re
from zipfile import BadZipFile, ZipFile
from xml.etree import ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from backend.domain.query_metric import *
HEADERS=("Запрос","Популярность запроса","Динамика за 28 дней","Динамика за 7 дней","Добавлений в корзину","Конверсия в корзину","Уникальные покупатели с заказами","Конверсия в заказ","Заказано на сумму по запросам, ₽","Запросы без действий","Доля запросов без действий")
HEADERS_V2=("Запрос","Популярность запроса","Динамика за 28 дней","Динамика за 7 дней","Добавлений в корзину","Конверсия в корзину","Уникальные покупатели с заказами","Конверсия в заказ","Заказано на сумму по запросам, ₽","Средняя цена","Показано товаров","Конкуренты","Запросы без действий","Доля запросов без действий","Запросы с похожими результатами","Доля запросов с похожими результатами","Запросы без результатов","Доля запросов без результатов")
SORT='Сортировка: По убыванию в Популярность запроса'
_BLANK=lambda x:x is None or x==''
_CELL=re.compile(r'([A-Z]+)([1-9]\d*)$')

def _raw(path):
 with ZipFile(path) as z:
  root=ET.fromstring(z.read('xl/workbook.xml'));ns={'m':'http://schemas.openxmlformats.org/spreadsheetml/2006/main','r':'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
  sheets=root.findall('m:sheets/m:sheet',ns)
  if len(sheets)!=1: return {},set()
  rid=sheets[0].get('{'+ns['r']+'}id');rels=ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
  target=None
  for rel in rels:
   if rel.get('Id')==rid:target=rel.get('Target');break
  if not target:raise ValueError('worksheet relationship missing')
  name=str(PurePosixPath('xl')/target) if not target.startswith('/') else target[1:]
  # relationship target often begins /xl or worksheets; normalize dot segments
  parts=[]
  for p in PurePosixPath(name).parts:
   if p=='..': parts.pop()
   elif p not in ('.','/'):parts.append(p)
  xml=ET.fromstring(z.read('/'.join(parts)));cells={};rows=set()
  for c in xml.iter('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
   coord=c.get('r');m=_CELL.fullmatch(coord or '')
   if not m:continue
   col=m.group(1);rn=int(m.group(2))
   if len(col)==1 and 'A'<=col<='R' and rn>=5:rows.add(rn)
   v=c.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v');cells[coord]=None if v is None else (v.text or '')
  return cells,rows

def _number(raw):
 if raw is None:raise ValueError
 d=Decimal(raw)
 if not d.is_finite():raise ValueError
 return d
def _native_number(cell, raw):
 # The OOXML lexical value preserves Decimal precision, while openpyxl's
 # metadata proves that the source cell was genuinely numeric.  In
 # particular, booleans use lexical 0/1 too and must not pass as numbers.
 if cell.data_type!='n' or isinstance(cell.value,bool) or cell.is_date:
  raise ValueError
 return _number(raw)
def _integer(raw):
 d=_number(raw)
 if d<0 or d!=d.to_integral_value():raise ValueError
 return int(d)
def _fraction(raw,upper=True):
 d=_number(raw)
 if d<0 or (upper and d>1):raise ValueError
 return d*100
def _query(v):
 if not isinstance(v,str):raise ValueError
 v=v.strip(' \u00a0')
 if not v:raise ValueError
 return v

def parse_ozon_query_metrics_xlsx(path:Path)->ParsedQueryMetricsReport:
 try:
  raw,candidates=_raw(path)
  with path.open('rb') as f:wb=load_workbook(f,data_only=False,read_only=False)
 except (BadZipFile, KeyError, ET.ParseError, InvalidFileException, OSError, ValueError) as e:
  raise QueryMetricsUnsupportedWorkbook('Не удалось прочитать XLSX.') from e
 try:
  if len(wb.worksheets)!=1:raise QueryMetricsIncompatibleReportSchema('Ожидается один лист.')
  ws=wb.active
  # Do not treat the package's frequently false ``A1`` dimension as business
  # coverage.  Exact V1/V2 signatures below still reject unknown shapes.
  ws.reset_dimensions() if hasattr(ws, "reset_dimensions") else None
  row5=tuple(ws.cell(5,c).value for c in range(1,33));row7=tuple(ws.cell(7,c).value for c in range(1,17));row6=tuple(ws.cell(6,c).value for c in range(1,12))
  if row5[0]=='Название товара' or row7[0]=='Позиция' or row6[0]=='SKU':raise QueryMetricsWrongReportType('Выбран другой тип отчёта.')
  if ws.merged_cells.ranges:raise QueryMetricsIncompatibleReportSchema('Объединённые ячейки не поддерживаются.')
  for r in range(1,5):
   for c in range(1,12):
    if ws.cell(r,c).data_type=='f':raise QueryMetricsIncompatibleReportSchema('Формулы в структуре отчёта не поддерживаются.')
  v1 = tuple(ws.cell(3,c).value for c in range(1,12))==HEADERS and ws['A2'].value==SORT
  v2 = (isinstance(ws['A2'].value,str) and ws['A2'].value.startswith('Поисковый запрос: ')
        and ws['A3'].value==SORT
        and tuple(ws.cell(4,c).value for c in range(1,19))==HEADERS_V2)
  v2_unfiltered = (ws['A2'].value==SORT
                   and tuple(ws.cell(3,c).value for c in range(1,19))==HEADERS_V2)
  if not (v1 or v2 or v2_unfiltered):raise QueryMetricsIncompatibleReportSchema('Структура отчёта изменилась.')
  width=11 if v1 else 18; data_start=6 if v2 else 5
  metadata_rows=(1,2,3) if v2 else (1,2)
  if any(not _BLANK(ws.cell(r,c).value) for r in metadata_rows for c in range(2,width+1)):raise QueryMetricsIncompatibleReportSchema('Структура метаданных изменилась.')
  for rn in range(1,max(ws.max_row,max(candidates,default=0))+1):
   for c in range(width+1,ws.max_column+1):
    if not _BLANK(ws.cell(rn,c).value):raise QueryMetricsIncompatibleReportSchema('Лишние бизнес-столбцы.')
  m=re.fullmatch(r'Период: (\d{2}\.\d{2}\.\d{4}) - (\d{2}\.\d{2}\.\d{4})',str(ws['A1'].value))
  try:
   if not m:raise ValueError
   start=datetime.strptime(m.group(1),'%d.%m.%Y').date();end=datetime.strptime(m.group(2),'%d.%m.%Y').date()
   if start>end:raise ValueError
  except ValueError as e:raise QueryMetricsInvalidReportPeriod('Некорректный период отчёта.') from e
  rows=[];errors=[];dupes=0;seen={};rows_seen=0
  messages=(('INVALID_QUERY','Некорректный поисковый запрос.'),('INVALID_POPULARITY','Некорректная популярность запроса.'),('INVALID_DYNAMICS','Некорректная динамика.'),('INVALID_DYNAMICS','Некорректная динамика.'),('INVALID_CART_ADD_USERS','Некорректное число добавлений в корзину.'),('INVALID_MARKET_CONVERSION','Некорректная рыночная конверсия.'),('INVALID_UNIQUE_BUYERS','Некорректное число покупателей.'),('INVALID_MARKET_CONVERSION','Некорректная рыночная конверсия.'),('INVALID_REVENUE','Некорректная сумма заказов.'),('INVALID_NO_ACTION_QUERIES','Некорректное число запросов без действий.'),('INVALID_NO_ACTION_SHARE','Некорректная доля запросов без действий.'))
  all_rows=sorted({r for r in candidates if r>=data_start}|{r for r in range(data_start,ws.max_row+1) if any(not _BLANK(ws.cell(r,c).value) for c in range(1,width+1))})
  for rn in all_rows:
   source_columns=range(1,12) if v1 else (1,2,3,4,5,6,7,8,9,13,14)
   cells=[ws.cell(rn,c) for c in source_columns]
   if all(_BLANK(c.value) for c in cells):continue
   rows_seen+=1;parsed=[];bad=None
   for i,c in enumerate(cells):
    try:
     if c.data_type=='f':raise ValueError
     coord=c.coordinate
     if i==0:x=_query(c.value)
     elif i in (1,4,6,9):x=_integer(str(_native_number(c,raw.get(coord))))
     elif i in (2,3):
      if c.data_type=='s' and c.value=='-':x=None
      else:x=_native_number(c,raw.get(coord))*100
     elif i in (5,7):x=_fraction(str(_native_number(c,raw.get(coord))))
     elif i==8:
      x=_native_number(c,raw.get(coord))
      if x<0:raise ValueError
     else:x=_fraction(str(_native_number(c,raw.get(coord))),upper=False)
     parsed.append(x)
    except (ValueError,InvalidOperation):bad=i;break
   if bad is not None:errors.append(QueryMetricRowError(rn,*messages[bad]));continue
   q=parsed[0];values=dict(zip(QUERY_METRIC_PAYLOAD_FIELDS,parsed[1:],strict=True));digest=query_metric_payload_sha256(values)
   if q in seen:
    if seen[q]!=digest:raise QueryMetricsConflictingObservationRows('Конфликтующие строки одного запроса.')
    dupes+=1;continue
   seen[q]=digest;rows.append(ParsedQueryMetricRow(rn,q,values,digest))
  return ParsedQueryMetricsReport(start,end,SORT,rows_seen,tuple(rows),tuple(errors),dupes,dupes)
 finally:wb.close()
