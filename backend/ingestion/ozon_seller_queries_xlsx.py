from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from backend.domain.product_query import *

HEADERS=("SKU","Артикул","Название товара","Запросы товара","Человек\nискало","Человек увидело","Позиция товара","Конверсия из\u00a0поиска в карточку","Конверсия из\u00a0поиска в заказ","Заказано товаров по\u00a0запросам","Заказано\u00a0на сумму\nпо\u00a0запросам")
_BLANK=lambda v:v is None or v==''
def _formula(c):return c.data_type=='f'
def _int(v):
 if not isinstance(v,str) or not re.fullmatch(r'\d+(?: \d+)*',v):raise ValueError
 return int(v.replace(' ',''))
def _pct(v):
 if not isinstance(v,str) or not re.fullmatch(r'\d+(?:,\d+)?%',v):raise ValueError
 d=Decimal(v[:-1].replace(',','.'))
 if d<0 or d>100:raise ValueError
 return d
def _money(v):
 if not isinstance(v,str) or not re.fullmatch(r'\d+(?: \d+)* ₽',v):raise ValueError
 return Decimal(v[:-2].replace(' ',''))
def _query(v):
 if not isinstance(v,str):raise ValueError
 v=v.strip(' \u00a0')
 if not v:raise ValueError
 return v

def parse_ozon_seller_queries_xlsx(path:Path)->ParsedSellerQueriesReport:
 try:
  with path.open('rb') as stream: wb=load_workbook(stream,data_only=False,read_only=False)
 except Exception as e: raise SellerQueriesUnsupportedWorkbook('Не удалось прочитать XLSX.') from e
 try:
  if len(wb.worksheets)!=1:raise SellerQueriesIncompatibleReportSchema('Ожидается один лист.')
  ws=wb.active
  # Clearly complete signatures belonging to known reports.
  row5=tuple(ws.cell(5,c).value for c in range(1,33));row7=tuple(ws.cell(7,c).value for c in range(1,17));row3=tuple(ws.cell(3,c).value for c in range(1,12))
  if row5[0]=='Название товара' or row7[0]=='Позиция' or row3[0]=='Запрос':raise SellerQueriesWrongReportType('Выбран другой тип отчёта.')
  structural=[ws.cell(r,c) for r in range(1,9) for c in range(1,12)]
  if any(_formula(c) for c in structural):raise SellerQueriesIncompatibleReportSchema('Формулы в структуре отчёта не поддерживаются.')
  if ws.merged_cells.ranges:raise SellerQueriesIncompatibleReportSchema('Объединённые ячейки не поддерживаются.')
  if tuple(ws.cell(6,c).value for c in range(1,12))!=HEADERS:raise SellerQueriesIncompatibleReportSchema('Структура заголовков отчёта изменена.')
  if any(not _BLANK(ws.cell(r,c).value) for r in (5,7) for c in range(1,12)):raise SellerQueriesIncompatibleReportSchema('Служебные строки отчёта изменены.')
  for r in range(1,5):
   if any(not _BLANK(ws.cell(r,c).value) for c in range(2,12)):raise SellerQueriesIncompatibleReportSchema('Структура метаданных изменена.')
  for r in range(1,max(ws.max_row,8)+1):
   for c in range(12,ws.max_column+1):
    if not _BLANK(ws.cell(r,c).value):raise SellerQueriesIncompatibleReportSchema('Лишние бизнес-столбцы.')
  try:
   dm=re.fullmatch(r'Дата: (\d{2}/\d{2}/\d{4})',str(ws['A1'].value));tm=re.fullmatch(r'Время: (\d{2}:\d{2}) \+00',str(ws['A2'].value))
   if not dm or not tm:raise ValueError
   generated=datetime.strptime(dm.group(1)+' '+tm.group(1),'%d/%m/%Y %H:%M').replace(tzinfo=timezone.utc)
  except ValueError as e:raise SellerQueriesInvalidGeneratedAt('Некорректное время формирования отчёта.') from e
  try:
   sm=re.fullmatch(r'Дата начала: (\d{2}/\d{2}/\d{4})',str(ws['A3'].value));em=re.fullmatch(r'Дата конца: (\d{2}/\d{2}/\d{4})',str(ws['A4'].value))
   if not sm or not em:raise ValueError
   start=datetime.strptime(sm.group(1),'%d/%m/%Y').date();end=datetime.strptime(em.group(1),'%d/%m/%Y').date()
   if start>end:raise ValueError
  except ValueError as e:raise SellerQueriesInvalidReportPeriod('Некорректный период отчёта.') from e
  rawid=ws['A8'].value
  if isinstance(rawid,bool) or not ((isinstance(rawid,int) and rawid>0) or (isinstance(rawid,str) and re.fullmatch(r'[1-9]\d*',rawid))):raise SellerQueriesInvalidProductContext('Некорректный SKU товара.')
  ozon=str(rawid)
  if not isinstance(ws['B8'].value,str) or not ws['B8'].value or not isinstance(ws['C8'].value,str) or not ws['C8'].value:raise SellerQueriesInvalidProductContext('Не указан контекст товара.')
  if any(not _BLANK(ws.cell(8,c).value) for c in range(4,12)):raise SellerQueriesIncompatibleReportSchema('Строка товара содержит лишние значения.')
  rows=[];errors=[];dupes=0;seen={};rows_seen=0
  messages=(('INVALID_QUERY','Некорректный поисковый запрос.'),('INVALID_SEARCHED_USERS','Некорректно указано количество искавших.'),('INVALID_SEEN_USERS','Некорректно указано количество увидевших.'),('INVALID_POSITION','Некорректная позиция товара.'),('INVALID_CONVERSION','Некорректное значение конверсии.'),('INVALID_CONVERSION','Некорректное значение конверсии.'),('INVALID_ORDERED_UNITS','Некорректно указано количество заказов.'),('INVALID_REVENUE','Некорректная сумма заказов.'))
  for rn in range(9,max(ws.max_row,8)+1):
   abc=[ws.cell(rn,c).value for c in range(1,4)]; cells=[ws.cell(rn,c) for c in range(4,12)]; vals=[c.value for c in cells]
   if all(_BLANK(v) for v in abc+vals):continue
   if any(not _BLANK(v) for v in abc):raise SellerQueriesIncompatibleReportSchema('Строка наблюдения нарушает схему.')
   rows_seen+=1
   funcs=(_query,_int,_int,_int,_pct,_pct,_int,_money);parsed=[];bad=None
   for i,(cell,fn) in enumerate(zip(cells,funcs)):
    try:
     if _formula(cell):raise ValueError
     parsed.append(fn(cell.value))
    except (ValueError,InvalidOperation):bad=i;break
   if bad is not None:
    errors.append(ProductQueryRowError(rn,*messages[bad]));continue
   q,searched,seen_users,pos,card,order,units,revenue=parsed
   values=dict(searched_users=searched,seen_users=seen_users,position_state=ProductQueryPositionState.SOURCE_ZERO if pos==0 else ProductQueryPositionState.KNOWN,average_position=None if pos==0 else pos,search_to_card_conversion_pct=card,search_to_order_conversion_pct=order,ordered_units=units,ordered_revenue_rub=revenue)
   digest=product_query_payload_sha256(values)
   if q in seen:
    if seen[q]!=digest:raise SellerQueriesConflictingObservationRows('Конфликтующие строки одного запроса.')
    dupes+=1;continue
   seen[q]=digest;rows.append(ParsedProductQueryRow(rn,q,values,digest))
  return ParsedSellerQueriesReport(generated,start,end,ozon,ws['B8'].value,ws['C8'].value,rows_seen,tuple(rows),tuple(errors),dupes,dupes)
 finally:wb.close()
