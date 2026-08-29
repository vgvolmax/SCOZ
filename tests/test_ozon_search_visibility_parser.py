from dataclasses import FrozenInstanceError, fields
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from time import perf_counter

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from backend.domain.search_visibility import (
    Cluster,
    CpcState,
    CpoState,
    OzonSearchVisibilityError,
    ParsedSearchVisibilityReport,
    ParsedSearchVisibilityRow,
    SEARCH_VISIBILITY_PAYLOAD_FIELDS,
    SearchQuery,
    SearchVisibilityConflictingObservationRows,
    SearchVisibilityIncompatibleReportSchema,
    SearchVisibilityInvalidObservedAt,
    SearchVisibilityInvalidSearchContext,
    SearchVisibilityUnsupportedWorkbook,
    SearchVisibilityWrongReportType,
    SearchVisibilityRowError,
    search_visibility_payload_sha256,
)
from backend.ingestion.ozon_search_visibility_xlsx import parse_ozon_search_visibility_xlsx
from tests.xlsx_factory import (
    OZON_SEARCH_VISIBILITY_HEADERS,
    build_ozon_products_workbook,
    build_ozon_search_visibility_workbook,
)


def _row(**changes: object) -> dict[str, object]:
    row = dict(zip(OZON_SEARCH_VISIBILITY_HEADERS, (
        "1147", 4218542117, "Синтетический товар", "Синтетический продавец",
        "0,052", "Продвигается", "22,32 ₽", "Средняя стоимость клика",
        "10%", "74,50", "4,8 (33 026 шт.)", "17 338 ₽", "2,60", "Да",
        "15-36 дней", "12,5%",
    ), strict=True))
    row.update(changes)
    return row


def _parse(tmp_path: Path, **workbook_options: object):
    path = tmp_path / "arbitrary.part"
    path.write_bytes(build_ozon_search_visibility_workbook(
        rows=(_row(),), **workbook_options
    ))
    return parse_ozon_search_visibility_xlsx(path)


def test_parses_exact_report_context_and_normalized_payload(tmp_path: Path) -> None:
    report = _parse(
        tmp_path,
        query=" \u00a0Смеситель  Для кухни\u00a0 ",
        cluster="\u00a0г. Москва, Россия ",
    )
    assert report.observed_at == datetime(2026, 8, 17, 3, 55, tzinfo=timezone.utc)
    assert report.query_text == "Смеситель  Для кухни"
    assert report.cluster_name == "г. Москва, Россия"
    assert (report.declared_rows, report.rows_seen) == (1, 1)
    values = report.rows[0].snapshot_values
    assert report.rows[0].ozon_product_id == "4218542117"
    assert values == {
        "source_title": "Синтетический товар", "seller_name": "Синтетический продавец",
        "position": 1147, "overall_score": Decimal("0.052"),
        "promotion_status": "Продвигается", "cpc_state": CpcState.ACTIVE,
        "cpc_rub": Decimal("22.32"),
        "promotion_strategy": "Средняя стоимость клика", "cpo_state": CpoState.ACTIVE,
        "cpo_pct": Decimal("10"), "relevance_score": Decimal("74.50"),
        "rating": Decimal("4.8"), "reviews_count": 33026,
        "buyer_price_rub": Decimal("17338"), "popularity_score": Decimal("2.60"),
        "ozon_promotion": True, "delivery_label": "15-36 дней",
        "delivery_min_days": 15, "delivery_max_days": 36,
        "price_index_pct": Decimal("12.5"),
    }


@pytest.mark.parametrize("field,value,expected", [
    ("Ставка\nОплата за заказ", "Выключено", CpoState.DISABLED),
    ("Ставка\nОплата за заказ", "—", CpoState.UNAVAILABLE),
])
def test_cpo_distinct_non_active_states(tmp_path: Path, field: str, value: str, expected: CpoState) -> None:
    report = _parse_rows(tmp_path, [_row(**{field: value})])
    assert report.rows[0].snapshot_values["cpo_state"] is expected
    assert report.rows[0].snapshot_values["cpo_pct"] is None

def test_cpc_numeric_and_disabled_are_distinct(tmp_path: Path) -> None:
    active=_parse_rows(tmp_path,[_row(**{"Ставка\nОплата за клик":"0,00 ₽"})]).rows[0]
    disabled=_parse_rows(tmp_path,[_row(**{"Ставка\nОплата за клик":"Выключено","Стратегия":"—"})]).rows[0]
    assert (active.snapshot_values['cpc_state'],active.snapshot_values['cpc_rub'])==(CpcState.ACTIVE,Decimal('0.00'))
    assert (disabled.snapshot_values['cpc_state'],disabled.snapshot_values['cpc_rub'])==(CpcState.DISABLED,None)
    assert active.payload_sha256 != disabled.payload_sha256

@pytest.mark.parametrize('sentinel',['—','Нет данных',''])
def test_unsupported_cpc_sentinel_is_rejected(tmp_path: Path, sentinel: str) -> None:
    report=_parse_rows(tmp_path,[_row(**{"Ставка\nОплата за клик":sentinel})])
    assert report.rows==() and report.row_errors[0].code=='INVALID_METRIC_VALUE'

def test_search_visibility_accepts_a1_stored_dimension_with_real_cells(tmp_path: Path) -> None:
    assert len(_parse(tmp_path,dimension_ref='A1').rows)==1


def test_search_visibility_parser_never_uses_random_read_only_cell_access(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(ReadOnlyWorksheet, "cell", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("random access")))
    report = _parse(tmp_path, dimension_ref="A1", q_z_values={"Z150": ""})
    assert len(report.rows) == 1


def test_search_visibility_realistic_report_parses_within_bound(tmp_path: Path) -> None:
    rows = []
    for index in range(150):
        rows.append(_row(**{"Позиция": str(index + 1), "ID товара": 4218542117 + index}))
    path = tmp_path / "realistic.xlsx"
    path.write_bytes(build_ozon_search_visibility_workbook(rows=rows, dimension_ref="A1", q_z_values={"Z159": ""}))
    started = perf_counter()
    report = parse_ozon_search_visibility_xlsx(path)
    elapsed = perf_counter() - started
    assert (report.declared_rows, report.rows_seen, len(report.rows)) == (150, 150, 150)
    assert report.rows[-1].ozon_product_id == str(4218542117 + 149)
    assert elapsed < 10


def test_exact_reviews_missing_sentinel_is_valid(tmp_path: Path) -> None:
    report = _parse_rows(tmp_path, [_row(Отзывы="— ")])
    assert report.rows[0].snapshot_values["rating"] is None
    assert report.rows[0].snapshot_values["reviews_count"] is None


def _parse_rows(tmp_path: Path, rows: list[dict[str, object]]):
    path = tmp_path / "rows.xlsx"
    path.write_bytes(build_ozon_search_visibility_workbook(rows=rows))
    return parse_ozon_search_visibility_xlsx(path)


@pytest.mark.parametrize("reviews", ["—", " — ", "", "-", "Нет данных", None])
def test_other_reviews_forms_are_recoverable_errors(tmp_path: Path, reviews: object) -> None:
    report = _parse_rows(tmp_path, [_row(Отзывы=reviews)])
    assert report.rows == ()
    assert report.row_errors[0].code == "INVALID_REVIEWS"
    assert report.row_errors[0].message == "Некорректное значение отзывов."


@pytest.mark.parametrize(("field", "value", "code"), [
    ("Позиция", "0", "INVALID_POSITION"),
    ("Позиция", 1, "INVALID_POSITION"),
    ("ID товара", True, "INVALID_PRODUCT_IDENTITY"),
    ("ID товара", 1.5, "INVALID_PRODUCT_IDENTITY"),
    ("ID товара", "1", "INVALID_PRODUCT_IDENTITY"),
    ("Ставка\nОплата за клик", "22,3 ₽", "INVALID_METRIC_VALUE"),
    ("Ставка\nОплата за заказ", "Нет данных", "INVALID_CPO_STATE"),
    ("Акции от Ozon", "да", "INVALID_METRIC_VALUE"),
    ("Срок доставки", "36-15 дней", "INVALID_DELIVERY"),
    ("Цена для покупателя", "1 999,00 ₽", "INVALID_METRIC_VALUE"),
])
def test_invalid_product_fields_are_ordered_recoverable_errors(
    tmp_path: Path, field: str, value: object, code: str
) -> None:
    report = _parse_rows(tmp_path, [_row(**{field: value})])
    assert report.rows == ()
    assert [(error.row, error.code) for error in report.row_errors] == [(10, code)]


def test_formula_in_product_row_is_recoverable(tmp_path: Path) -> None:
    report = _parse(tmp_path, formula_cells={"E10": "=1+1"})
    assert report.rows == ()
    assert report.row_errors[0].code == "INVALID_METRIC_VALUE"


@pytest.mark.parametrize("options", [
    {"row_6_values": {1: " "}}, {"row_8_values": {1: "\u00a0"}},
    {"formula_cells": {"A9": "=1+1"}}, {"extra_sheet": True},
    {"merged_cells": ("A1:B1",)}, {"q_z_values": {"Q10": "business"}},
])
def test_incompatible_structural_variants(tmp_path: Path, options: dict[str, object]) -> None:
    with pytest.raises(SearchVisibilityIncompatibleReportSchema):
        _parse(tmp_path, **options)


def test_row_9_ordinary_text_and_empty_q_z_are_accepted(tmp_path: Path) -> None:
    report = _parse(tmp_path, row_9_values={1: "любое пояснение", 16: "текст"}, q_z_values={"Z1000": ""})
    assert len(report.rows) == 1


def test_exact_header_newline_is_required(tmp_path: Path) -> None:
    headers = list(OZON_SEARCH_VISIBILITY_HEADERS)
    headers[6] = "Ставка Оплата за клик"
    with pytest.raises(SearchVisibilityIncompatibleReportSchema):
        _parse(tmp_path, headers=headers)


@pytest.mark.parametrize(("date", "time"), [
    ("31/02/2026", "03:55 +00"), ("17/08/2026", "24:00 +00"),
    ("2026-08-17", "03:55 +00"), ("17/08/2026", "03:55 +03"),
])
def test_invalid_observation_time_is_fatal(tmp_path: Path, date: str, time: str) -> None:
    with pytest.raises(SearchVisibilityInvalidObservedAt):
        _parse(tmp_path, date=date, time=time)


@pytest.mark.parametrize("identity", ["", " ", "\u00a0", " \u00a0 "])
def test_empty_query_after_exact_edge_cleanup_is_fatal(tmp_path: Path, identity: str) -> None:
    with pytest.raises(SearchVisibilityInvalidSearchContext):
        _parse(tmp_path, query=identity)


def test_declared_candidate_mismatch_is_fatal(tmp_path: Path) -> None:
    with pytest.raises(SearchVisibilityIncompatibleReportSchema):
        _parse(tmp_path, declared_rows=2)


def test_identical_duplicate_keeps_first_and_counts_warning(tmp_path: Path) -> None:
    report = _parse_rows(tmp_path, [_row(), _row()])
    assert [row.source_row for row in report.rows] == [10]
    assert (report.duplicate_input_rows, report.warnings_count) == (1, 1)


def test_conflicting_product_duplicate_is_fatal(tmp_path: Path) -> None:
    with pytest.raises(SearchVisibilityConflictingObservationRows):
        _parse_rows(tmp_path, [_row(), _row(**{"Позиция": "2"})])


def test_unreadable_package_is_unsupported(tmp_path: Path) -> None:
    path = tmp_path / "bad.part"
    path.write_bytes(b"not xlsx")
    with pytest.raises(SearchVisibilityUnsupportedWorkbook):
        parse_ozon_search_visibility_xlsx(path)


def test_products_workbook_is_wrong_report_type(tmp_path: Path) -> None:
    path = tmp_path / "products.xlsx"
    path.write_bytes(build_ozon_products_workbook())
    with pytest.raises(SearchVisibilityWrongReportType):
        parse_ozon_search_visibility_xlsx(path)


def test_seller_queries_workbook_is_wrong_report_type(tmp_path: Path) -> None:
    path = tmp_path / "seller-queries.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    for row, value in enumerate((
        "Дата: 17/08/2026",
        "Время: 03:56 +00",
        "Дата начала: 18/07/2026",
        "Дата конца: 14/08/2026",
    ), start=1):
        sheet.cell(row=row, column=1, value=value)
    for column, value in enumerate(
        ("SKU", "Артикул", "Название товара", "Запросы товара"), start=1
    ):
        sheet.cell(row=6, column=column, value=value)
    workbook.save(path)
    workbook.close()

    with pytest.raises(SearchVisibilityWrongReportType):
        parse_ozon_search_visibility_xlsx(path)


def test_unrelated_workbook_is_wrong_report_type(tmp_path: Path) -> None:
    path = tmp_path / "unrelated.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "unrelated readable content"
    workbook.save(path)
    workbook.close()

    with pytest.raises(SearchVisibilityWrongReportType):
        parse_ozon_search_visibility_xlsx(path)


def test_partial_metadata_markers_are_incompatible_schema(tmp_path: Path) -> None:
    path = tmp_path / "damaged-explainer.xlsx"
    path.write_bytes(build_ozon_search_visibility_workbook())
    workbook = load_workbook(path)
    workbook.active["A1"] = "damaged marker"
    workbook.save(path)
    workbook.close()
    with pytest.raises(SearchVisibilityIncompatibleReportSchema):
        parse_ozon_search_visibility_xlsx(path)


def test_frozen_domain_contract_field_order_and_error_hierarchy() -> None:
    assert [field.name for field in fields(SearchQuery)] == ["id", "query_text", "created_at"]
    assert [field.name for field in fields(Cluster)] == ["id", "name", "created_at"]
    assert [field.name for field in fields(ParsedSearchVisibilityRow)] == [
        "source_row", "ozon_product_id", "snapshot_values", "payload_sha256",
    ]
    assert [field.name for field in fields(ParsedSearchVisibilityReport)] == [
        "observed_at", "query_text", "cluster_name", "declared_rows", "rows_seen",
        "rows", "row_errors", "duplicate_input_rows", "warnings_count",
    ]
    with pytest.raises(FrozenInstanceError):
        SearchVisibilityRowError(10, "X", "x").row = 11
    module = __import__("backend.domain.search_visibility", fromlist=["ignored"])
    assert set(OzonSearchVisibilityError.__subclasses__()) == {
        getattr(module, name) for name in (
            "SearchVisibilityUnsupportedWorkbook", "SearchVisibilityWrongReportType",
            "SearchVisibilityIncompatibleReportSchema", "SearchVisibilityInvalidObservedAt",
            "SearchVisibilityInvalidSearchContext", "SearchVisibilityInvalidProductIdentity",
            "SearchVisibilityInvalidMetricValue", "SearchVisibilityConflictingObservationRows",
            "SearchVisibilityNoUsableRows", "SearchVisibilityConcurrentImportConflict",
            "SearchVisibilityUploadTooLarge", "SearchVisibilityUnsupportedUploadMediaType",
            "SearchVisibilityImportPersistenceError",
        )
    }


def test_exact_payload_order_and_canonical_decimal_cpo_hashing(tmp_path: Path) -> None:
    assert SEARCH_VISIBILITY_PAYLOAD_FIELDS == (
        "source_title", "seller_name", "position", "overall_score", "promotion_status",
        "cpc_state", "cpc_rub", "promotion_strategy", "cpo_state", "cpo_pct", "relevance_score",
        "rating", "reviews_count", "buyer_price_rub", "popularity_score",
        "ozon_promotion", "delivery_label", "delivery_min_days", "delivery_max_days",
        "price_index_pct",
    )
    values = _parse(tmp_path).rows[0].snapshot_values
    equivalent = dict(values, overall_score=Decimal("0.0520"), cpo_state=CpoState("ACTIVE"))
    assert search_visibility_payload_sha256(values) == search_visibility_payload_sha256(equivalent)
    changed = dict(values, cpo_state=CpoState.DISABLED, cpo_pct=None)
    assert search_visibility_payload_sha256(values) != search_visibility_payload_sha256(changed)
