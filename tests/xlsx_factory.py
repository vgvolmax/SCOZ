from io import BytesIO
from typing import Mapping, Sequence
from zipfile import ZIP_DEFLATED, ZipFile
import re

from openpyxl import Workbook

OZON_SELLER_QUERIES_HEADERS = ("SKU", "Артикул", "Название товара", "Запросы товара", "Человек\nискало", "Человек увидело", "Позиция товара", "Конверсия из\u00a0поиска в карточку", "Конверсия из\u00a0поиска в заказ", "Заказано товаров по\u00a0запросам", "Заказано\u00a0на сумму\nпо\u00a0запросам")
OZON_QUERY_METRICS_HEADERS = ("Запрос", "Популярность запроса", "Динамика за 28 дней", "Динамика за 7 дней", "Добавлений в корзину", "Конверсия в корзину", "Уникальные покупатели с заказами", "Конверсия в заказ", "Заказано на сумму по запросам, ₽", "Запросы без действий", "Доля запросов без действий")
OZON_QUERY_METRICS_V2_HEADERS = ("Запрос", "Популярность запроса", "Динамика за 28 дней", "Динамика за 7 дней", "Добавлений в корзину", "Конверсия в корзину", "Уникальные покупатели с заказами", "Конверсия в заказ", "Заказано на сумму по запросам, ₽", "Средняя цена", "Показано товаров", "Конкуренты", "Запросы без действий", "Доля запросов без действий", "Запросы с похожими результатами", "Доля запросов с похожими результатами", "Запросы без результатов", "Доля запросов без результатов")

def _patch_dimension(data: bytes, ref: str | None) -> bytes:
    if ref is None: return data
    source, target = BytesIO(data), BytesIO()
    with ZipFile(source) as zin, ZipFile(target, "w", ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            body=zin.read(info.filename)
            if info.filename=="xl/worksheets/sheet1.xml":
                body=re.sub(rb'<dimension ref="[^"]+"', f'<dimension ref="{ref}"'.encode(), body, count=1)
            zout.writestr(info,body)
    return target.getvalue()

def _save(workbook: Workbook) -> bytes:
    output = BytesIO(); workbook.save(output); workbook.close(); return output.getvalue()

def build_ozon_seller_queries_workbook(*, rows: Sequence[Mapping[str, object]] | None = None,
        headers: Sequence[str] = OZON_SELLER_QUERIES_HEADERS, extra_sheet: bool = False,
        merged_cells: Sequence[str] = (), formula_cells: Mapping[str, str] | None = None,
        l_plus_values: Mapping[str, object] | None = None,
        date: str = "18/08/2026", time: str = "04:10 +00", period_start: str = "20/07/2026",
        period_end: str = "17/08/2026", ozon_id: object = 100000001,
        article: object = "SYNTH-001", title: object = "Синтетический товар") -> bytes:
    wb = Workbook(); ws = wb.active
    ws["A1"] = f"Дата: {date}"; ws["A2"] = f"Время: {time}"
    ws["A3"] = f"Дата начала: {period_start}"; ws["A4"] = f"Дата конца: {period_end}"
    for col, header in enumerate(headers, 1): ws.cell(6, col, header)
    ws.cell(8, 1, ozon_id); ws.cell(8, 2, article); ws.cell(8, 3, title)
    source = rows if rows is not None else ({headers[3]: "синтетический запрос", headers[4]: "1 000", headers[5]: "900", headers[6]: "1", headers[7]: "10%", headers[8]: "2%", headers[9]: "20", headers[10]: "5 000 ₽"},)
    for rn, row in enumerate(source, 9):
        for col, header in enumerate(headers, 1): ws.cell(rn, col, row.get(header))
    for cell, value in (formula_cells or {}).items(): ws[cell] = value
    for cell, value in (l_plus_values or {}).items(): ws[cell] = value
    for area in merged_cells: ws.merge_cells(area)
    if extra_sheet: wb.create_sheet()
    return _save(wb)

def build_ozon_query_metrics_workbook(*, rows: Sequence[Mapping[str, object]] | None = None,
        headers: Sequence[str] = OZON_QUERY_METRICS_HEADERS, period: str = "21.07.2026 - 17.08.2026",
        sort_context: str = "Сортировка: По убыванию в Популярность запроса",
        extra_sheet: bool = False, merged_cells: Sequence[str] = (),
        formula_cells: Mapping[str, str] | None = None, l_plus_values: Mapping[str, object] | None = None,
        raw_numeric_overrides: Mapping[str, str] | None = None, dimension_ref: str | None = None,
        horizontal_capitalized: bool = False) -> bytes:
    wb = Workbook(); ws = wb.active; ws["A1"] = f"Период: {period}"; ws["A2"] = sort_context
    for col, header in enumerate(headers, 1): ws.cell(3, col, header)
    ws["A4"] = "—"
    source = rows if rows is not None else ({headers[0]: "синтетический запрос", headers[1]: 1000, headers[2]: 0.1, headers[3]: "-", headers[4]: 100, headers[5]: 0.1, headers[6]: 50, headers[7]: 0.05, headers[8]: 1234.5, headers[9]: 200, headers[10]: 0.2},)
    for rn, row in enumerate(source, 5):
        for col, header in enumerate(headers, 1): ws.cell(rn, col, row.get(header))
    for cell, value in (formula_cells or {}).items(): ws[cell] = value
    for cell, value in (l_plus_values or {}).items(): ws[cell] = value
    for area in merged_cells: ws.merge_cells(area)
    if horizontal_capitalized:
        from openpyxl.styles import Alignment
        ws["A1"].alignment = Alignment(horizontal="left")
    if extra_sheet: wb.create_sheet()
    original = _save(wb)
    if not (raw_numeric_overrides or dimension_ref or horizontal_capitalized): return original
    source_io, target = BytesIO(original), BytesIO()
    with ZipFile(source_io) as zin, ZipFile(target, "w", ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                text = data.decode()
                if dimension_ref: text = re.sub(r'<dimension ref="[^"]+"', f'<dimension ref="{dimension_ref}"', text, count=1)
                for coord, raw in (raw_numeric_overrides or {}).items():
                    pattern = rf'(<c[^>]*\br="{re.escape(coord)}"[^>]*>.*?<v>)[^<]*(</v>)'
                    text, count = re.subn(pattern, rf'\g<1>{raw}\g<2>', text, count=1)
                    if not count: raise ValueError(f"cell {coord} has no numeric value")
                data = text.encode()
            elif info.filename == "xl/styles.xml" and horizontal_capitalized:
                data = data.replace(b'horizontal="left"', b'horizontal="Left"').replace(b'horizontal="right"', b'horizontal="Right"')
            zout.writestr(info, data)
    return target.getvalue()

def build_ozon_query_metrics_v2_workbook(*, filter_text="герметик", headers=OZON_QUERY_METRICS_V2_HEADERS,
        rows: Sequence[Mapping[str, object]] | None=None, dimension_ref: str|None=None) -> bytes:
    wb=Workbook();ws=wb.active
    ws['A1']='Период: 21.07.2026 - 17.08.2026';ws['A2']=f'Поисковый запрос: {filter_text}'
    ws['A3']='Сортировка: По убыванию в Популярность запроса'
    for col,header in enumerate(headers,1):ws.cell(4,col,header)
    ws['A5']='—'
    source=rows if rows is not None else ({headers[0]:'синтетический запрос',headers[1]:1000,headers[2]:.1,headers[3]:'-',headers[4]:100,headers[5]:.1,headers[6]:50,headers[7]:.05,headers[8]:1234.5,headers[9]:100,headers[10]:25,headers[11]:10,headers[12]:200,headers[13]:.2,headers[14]:30,headers[15]:.03,headers[16]:4,headers[17]:.004},)
    for rn,row in enumerate(source,6):
        for col,header in enumerate(headers,1):ws.cell(rn,col,row.get(header))
    return _patch_dimension(_save(wb),dimension_ref)

def build_ozon_query_metrics_v2_unfiltered_workbook(*, headers=OZON_QUERY_METRICS_V2_HEADERS,
        rows: Sequence[Mapping[str, object]] | None=None, dimension_ref: str|None=None,
        sort_context: str='Сортировка: По убыванию в Популярность запроса') -> bytes:
    wb=Workbook();ws=wb.active
    ws['A1']='Период: 21.07.2026 - 17.08.2026'
    ws['A2']=sort_context
    for col,header in enumerate(headers,1):ws.cell(3,col,header)
    ws['A4']='—'
    source=rows if rows is not None else ({headers[0]:'синтетический запрос',headers[1]:1000,headers[2]:.1,headers[3]:'-',headers[4]:100,headers[5]:.1,headers[6]:50,headers[7]:.05,headers[8]:1234.5,headers[9]:100,headers[10]:25,headers[11]:10,headers[12]:200,headers[13]:.2,headers[14]:30,headers[15]:.03,headers[16]:4,headers[17]:.004},)
    for rn,row in enumerate(source,5):
        for col,header in enumerate(headers,1):ws.cell(rn,col,row.get(header))
    return _patch_dimension(_save(wb),dimension_ref)


OZON_PRODUCTS_HEADERS = (
    "Название товара",
    "Ссылка на товар",
    "Продавец",
    "Бренд",
    "Категория 1 уровня",
    "Категория 3 уровня",
    "Признак товара",
    "Заказано на сумму, ₽",
    "Динамика оборота, %",
    "Заказано, штуки",
    "Средняя цена, ₽",
    "Минимальная цена, ₽",
    "Доля выкупа, %",
    "Упущенные продажи",
    "Дней без остатка",
    "Среднесуточные продажи, ₽",
    "Среднесуточные продажи, штуки",
    "Остаток на конец периода, штуки",
    "Схема работы",
    "Объем товара, л",
    "Показы всего",
    "Просмотры в поиске и каталоге",
    "Просмотры карточки",
    "Конверсия из показа в заказ, %",
    "В корзину из поиска и каталога, %",
    "В корзину из карточки, %",
    "Скидка за счет акций",
    "Доля суммы заказов по акциям, %",
    "Дней в акциях",
    "Дней с продвижением",
    "Доля рекламных расходов, %",
    "Дата создания карточки товара",
)

OZON_SEARCH_VISIBILITY_HEADERS = (
    "Позиция", "ID товара", "Название товара", "Имя селлера", "Сводная оценка",
    "Статус", "Ставка\nОплата за клик", "Стратегия",
    "Ставка\nОплата за заказ", "Соответствие запросу", "Отзывы",
    "Цена для покупателя", "Популярность общая", "Акции от Ozon",
    "Срок доставки", "Индекс цен",
)


def _default_search_visibility_row() -> dict[str, object]:
    return dict(zip(OZON_SEARCH_VISIBILITY_HEADERS, (
        "1", 100000001, "Синтетический товар", "Синтетический продавец", "0,526",
        "Продвигается", "10,50 ₽", "Автостратегия", "5%", "99,1",
        "4,8 (1 234 шт.)", "1 999 ₽", "42,2", "Да", "1-2 дня", "10,0%",
    ), strict=True))


def build_ozon_search_visibility_workbook(
    *, query: str = "тестовый запрос", cluster: str = "г. Тестоград, Россия",
    date: str = "17/08/2026", time: str = "03:55 +00",
    declared_rows: int | None = None,
    rows: Sequence[Mapping[str, object]] | None = None,
    headers: Sequence[str] = OZON_SEARCH_VISIBILITY_HEADERS,
    extra_sheet: bool = False, merged_cells: Sequence[str] = (),
    formula_cells: Mapping[str, str] | None = None,
    q_z_values: Mapping[str, object] | None = None,
    row_6_values: Mapping[int, object] | None = None,
    row_8_values: Mapping[int, object] | None = None,
    row_9_values: Mapping[int, object] | None = None,
    dimension_ref: str | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    source_rows = rows if rows is not None else (_default_search_visibility_row(),)
    declared = len(source_rows) if declared_rows is None else declared_rows
    for coordinate, value in {
        "A1": f"Дата: {date}", "A2": f"Запрос: {query}",
        "A3": f"Время: {time}", "A4": f"Регион: {cluster}",
        "A5": f"Сколько позиций в выдаче: {declared}",
    }.items():
        sheet[coordinate] = value
    for column, header in enumerate(headers, 1):
        sheet.cell(7, column, header)
    for column, value in (row_6_values or {}).items(): sheet.cell(6, column, value)
    for column, value in (row_8_values or {}).items(): sheet.cell(8, column, value)
    help_values = row_9_values or {1: "Синтетическая поясняющая строка"}
    for column, value in help_values.items(): sheet.cell(9, column, value)
    for row_number, row in enumerate(source_rows, 10):
        for column, header in enumerate(headers, 1):
            value = row.get(header)
            # openpyxl serializes 1.0 as the integer lexical form ``1``. Preserve
            # this deliberately invalid mutation so parser tests can distinguish it.
            if header == "ID товара" and isinstance(value, float):
                value = str(value)
            sheet.cell(row_number, column, value)
    for coordinate, value in (q_z_values or {}).items(): sheet[coordinate] = value
    for coordinate, value in (formula_cells or {}).items(): sheet[coordinate] = value
    for cell_range in merged_cells: sheet.merge_cells(cell_range)
    if extra_sheet: workbook.create_sheet()
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return _patch_dimension(output.getvalue(), dimension_ref)


def _default_row(category_level_3: str) -> dict[str, object]:
    return dict(zip(OZON_PRODUCTS_HEADERS, (
        "Синтетический товар",
        "https://www.ozon.ru/product/100000001/",
        "Синтетический продавец",
        "Синтетический бренд",
        "Синтетическая категория 1",
        category_level_3,
        None,
        1000,
        1.31,
        2,
        500,
        450,
        95.8,
        0,
        "0 из 28",
        0,
        0,
        0,
        "FBO",
        1.5,
        100,
        50,
        25,
        2.5,
        5.0,
        10.0,
        0,
        0,
        "0 из 7",
        "0 из 7",
        7.7,
        "2026-01-01",
    ), strict=True))


def build_ozon_products_workbook(
    *,
    rows: Sequence[Mapping[str, object]] | None = None,
    generated_on: str = "08.16.26",
    window_label: str = "7 дней",
    category_level_3: str = "Синтетическая категория",
    marker_overrides: Mapping[str, object] | None = None,
    headers: Sequence[str] = OZON_PRODUCTS_HEADERS,
    extra_sheet: bool = False,
    dimension_ref: str | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    markers: dict[str, object] = {
        "A1": "Дата формирования:",
        "B1": generated_on,
        "A2": "Период отчета:",
        "B2": window_label,
        "A3": "Категория 3 уровня:",
        "B3": category_level_3,
        "A6": "Среднее значение по товарам",
    }
    markers.update(marker_overrides or {})
    for coordinate, value in markers.items():
        sheet[coordinate] = value
    sheet.append([])  # Rows 1-3 already exist; preserve blank contract row 4.
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=5, column=column, value=header)
    source_rows = rows if rows is not None else (_default_row(category_level_3),)
    for row_number, row in enumerate(source_rows, start=7):
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=row_number, column=column, value=row.get(header))
    if extra_sheet:
        workbook.create_sheet()
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return _patch_dimension(output.getvalue(), dimension_ref)
