from pathlib import Path
from io import BytesIO

from backend.config import DATA_DIR, DEFAULT_DB_PATH


ROOT = Path(__file__).resolve().parents[1]


def test_runtime_inputs():
    assert (ROOT / "VERSION.txt").read_text(
        encoding="utf-8"
    ).strip() == "0.1.0"
    assert (ROOT / "requirements.txt").read_text().splitlines() == [
        "fastapi==0.139.2",
        "uvicorn==0.51.0",
        "openpyxl==3.1.5",
        "python-multipart==0.0.32",
        "httpx==0.28.1",
    ]
    assert "pytest" in (ROOT / "requirements-dev.txt").read_text()
    assert DEFAULT_DB_PATH == DATA_DIR / "scoz.db"
    assert DEFAULT_DB_PATH.parent.name == "data"
    assert "runtime" not in DEFAULT_DB_PATH.parts


def test_runtime_inputs_include_exact_pr3_dependencies_and_no_pandas():
    requirements = (ROOT / "requirements.txt").read_text().splitlines()
    assert requirements == [
        "fastapi==0.139.2",
        "uvicorn==0.51.0",
        "openpyxl==3.1.5",
        "python-multipart==0.0.32",
        "httpx==0.28.1",
    ]
    assert all("pandas" not in requirement.lower() for requirement in requirements)


def test_bootstrap_contract():
    text = (ROOT / "start.bat").read_text(encoding="utf-8")
    required = [
        "chcp 65001 >nul", 'set "PYTHONUTF8=1"', 'set "PYTHONIOENCODING=utf-8"',
        'cd /d "%~dp0"', "python-3.13.14-embed-amd64.zip", "https://www.python.org/ftp/python/3.13.14/",
        "https://bootstrap.pypa.io/get-pip.py", ".part", "Entries.Count", "5000000", "100000",
        "python313.zip", "Lib\\site-packages", "import site", '-m pip install -r "requirements.txt"',
        "(3,13,14)", "AMD64", "fastapi", "0.139.2", "uvicorn", "0.51.0", 'rmdir /s /q "runtime"',
        '"runtime\\python.exe" "launcher.py"',
    ]
    for value in required:
        assert value in text
    lowered = text.lower()
    assert "npm " not in lowered and "node " not in lowered and "frontend build" not in lowered
    assert 'rmdir /s /q "data"' not in lowered
    assert 'rmdir /s /q "scoz.db"' not in lowered
    assert 'del /q "data' not in lowered
    assert 'del /q "scoz.db"' not in lowered


def test_bootstrap_validates_pr3_imports_and_distribution_versions():
    text = (ROOT / "start.bat").read_text(encoding="utf-8")
    assert "import fastapi,uvicorn,openpyxl,multipart" in text
    assert "m.version('openpyxl') == '3.1.5'" in text
    assert "m.version('python-multipart') == '0.0.32'" in text


def test_windows_smoke_covers_pr4_portable_persistence_and_recovery():
    text = (ROOT / "tests/windows_smoke.ps1").read_text(encoding="ascii")
    for required in (
        "(3, 'ozon_search_visibility_import')",
        "build_ozon_search_visibility_workbook",
        "parse_ozon_search_visibility_xlsx",
        "search_visibility_snapshots",
        "ozon_products_xlsx",
        "ozon_search_visibility_xlsx",
        "recover_interrupted_ozon_products_imports",
        "recover_interrupted_ozon_search_visibility_imports",
        "Portable imports or archives did not survive",
    ):
        assert required in text


def test_strict_xlsx_factory_emits_contract_rows_in_memory():
    from openpyxl import load_workbook

    from tests.xlsx_factory import OZON_PRODUCTS_HEADERS, build_ozon_products_workbook

    existing_xlsx = set(ROOT.rglob("*.xlsx"))
    data = build_ozon_products_workbook()
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=False)
    assert len(workbook.sheetnames) == 1
    sheet = workbook.active
    assert tuple(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0][:2] == (
        "Дата формирования:", "08.16.26"
    )
    assert tuple(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0][:2] == (
        "Период отчета:", "7 дней"
    )
    assert tuple(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0][:2] == (
        "Категория 3 уровня:", "Синтетическая категория"
    )
    assert all(value is None for value in next(sheet.iter_rows(min_row=4, max_row=4, max_col=32, values_only=True)))
    assert tuple(next(sheet.iter_rows(min_row=5, max_row=5, max_col=32, values_only=True))) == OZON_PRODUCTS_HEADERS
    summary = next(sheet.iter_rows(min_row=6, max_row=6, max_col=32, values_only=True))
    assert summary[0] == "Среднее значение по товарам"
    assert all(value is None for value in summary[1:])
    row = next(sheet.iter_rows(min_row=7, max_row=7, max_col=32, values_only=True))
    assert row[1] == "https://www.ozon.ru/product/100000001/"
    workbook.close()
    assert set(ROOT.rglob("*.xlsx")) == existing_xlsx


def test_readme_documents_user_owned_database_lifecycle():
    text = (ROOT / "README.md").read_text(encoding="utf-8")
    assert "`data/scoz.db`" in text
    assert "user-owned persistent SQLite" in text
    assert "pending schema migrations apply automatically before a new local server starts" in text


def test_ignored_generated_and_sensitive_state():
    ignore = (ROOT / ".gitignore").read_text()
    for item in ("runtime/", "data/", ".venv/", "__pycache__/", "*.enc.json", "*credentials*.json"):
        assert item in ignore
    wrapper = (ROOT / "RUN_SERVER.cmd").read_text()
    assert '"%~dp0"' in wrapper
    assert "server.pid" in wrapper and "server_console.log" in wrapper
