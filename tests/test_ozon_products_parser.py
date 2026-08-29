from dataclasses import fields
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from time import perf_counter

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from backend.domain.product_snapshot import (
    ConflictingObservationRows, IncompatibleReportSchema, InvalidMetricValue, InvalidReportPeriod,
    OzonProductsImportSummary, ParsedOzonProductRow, ParsedOzonProductsReport,
    ProductSnapshot, RowError, SnapshotWriteKind, SnapshotWriteResult, WrongReportType,
    canonical_decimal_text, decimal_from_excel_number, product_snapshot_payload,
)
from backend.domain.lineage import normalized_payload_sha256
from backend.ingestion.ozon_products_xlsx import parse_ozon_products_xlsx
from tests.xlsx_factory import (OZON_PRODUCTS_HEADERS, _default_row, build_ozon_products_workbook)


def test_frozen_domain_field_counts() -> None:
    assert len(fields(ProductSnapshot)) == 45
    assert len(fields(OzonProductsImportSummary)) == 17


def test_canonical_decimal_text() -> None:
    assert canonical_decimal_text(Decimal("1.2300")) == "1.23"
    assert canonical_decimal_text(Decimal("-0.00")) == "0"

PRODUCT_SNAPSHOT_FIELDS = [
    "id", "product_id", "report_generated_on", "report_window_days", "revision",
    "supersedes_snapshot_id", "payload_sha256", "import_batch_id", "source_artifact_id",
    "imported_at", "product_url", "title", "seller_name", "brand", "category_level_1",
    "category_level_3", "product_badges", "ordered_amount_rub", "turnover_change_pct",
    "ordered_units", "average_price_rub", "minimum_price_rub", "buyout_share_pct",
    "missed_sales_source_value", "out_of_stock_days", "out_of_stock_window_days",
    "avg_daily_sales_rub", "avg_daily_sales_units", "stock_end_units", "fulfillment_scheme",
    "volume_l", "impressions_total", "search_catalog_views", "card_views",
    "impression_to_order_pct", "search_catalog_to_cart_pct", "card_to_cart_pct",
    "promotion_discount_source_value", "promotion_order_amount_share_pct", "promotion_days",
    "promotion_window_days", "advertising_days", "advertising_window_days", "total_drr_pct",
    "card_created_on",
]
SUMMARY_FIELDS = [
    "import_batch_id", "source", "import_kind", "status", "report_generated_on",
    "report_window_days", "rows_seen", "rows_accepted", "rows_skipped",
    "duplicate_observations", "new_observations", "corrected_revisions", "warnings_count",
    "row_errors_total", "started_at", "finished_at", "source_artifact",
]


def _parse_bytes(tmp_path, data):
    path = tmp_path / "arbitrary-name.xlsx"
    path.write_bytes(data)
    return parse_ozon_products_xlsx(path)


def test_exact_frozen_domain_interfaces():
    assert [field.name for field in fields(ProductSnapshot)] == PRODUCT_SNAPSHOT_FIELDS
    assert [field.name for field in fields(OzonProductsImportSummary)] == SUMMARY_FIELDS
    assert [field.name for field in fields(ParsedOzonProductRow)] == ["source_row", "ozon_product_id", "snapshot_values", "payload_sha256"]
    assert [field.name for field in fields(RowError)] == ["row", "code", "message"]
    assert [field.name for field in fields(ParsedOzonProductsReport)] == ["report_generated_on", "report_window_days", "rows_seen", "rows", "row_errors", "duplicate_input_rows", "warnings_count"]
    assert [item.value for item in SnapshotWriteKind] == ["NEW", "DUPLICATE", "CORRECTED"]
    assert [field.name for field in fields(SnapshotWriteResult)] == ["kind", "snapshot"]


@pytest.mark.parametrize("value, expected", [(1, Decimal("1")), (1.31, Decimal("1.31"))])
def test_decimal_excel_numbers(value, expected):
    assert decimal_from_excel_number(value) == expected


@pytest.mark.parametrize("value", [True, "1,31", float("nan"), float("inf"), float("-inf")])
def test_decimal_excel_numbers_reject_invalid(value):
    with pytest.raises(InvalidMetricValue):
        decimal_from_excel_number(value)


def test_payload_contract_and_stable_hash(tmp_path):
    row = _parse_bytes(tmp_path, build_ozon_products_workbook()).rows[0]
    payload = product_snapshot_payload(row.snapshot_values)
    assert list(payload) == PRODUCT_SNAPSHOT_FIELDS[10:]
    assert payload["turnover_change_pct"] == "1.31"
    assert payload["product_badges"] is None and payload["ordered_units"] == 2
    assert payload["card_created_on"] == "2026-01-01"
    assert normalized_payload_sha256(payload) == normalized_payload_sha256(dict(reversed(list(payload.items()))))
    for changed in ({k: v for k, v in row.snapshot_values.items() if k != "title"}, {**row.snapshot_values, "extra": 1}):
        with pytest.raises(ValueError):
            product_snapshot_payload(changed)


@pytest.mark.parametrize("kwargs,error", [
    ({"extra_sheet": True}, IncompatibleReportSchema),
    ({"headers": tuple(reversed(OZON_PRODUCTS_HEADERS))}, IncompatibleReportSchema),
    ({"generated_on": "2026-08-16"}, InvalidReportPeriod),
    ({"window_label": "7 days"}, InvalidReportPeriod),
    ({"marker_overrides": {"A1": "Другой отчёт"}}, WrongReportType),
    ({"marker_overrides": {"A6": "wrong"}}, IncompatibleReportSchema),
    ({"marker_overrides": {"B3": None}}, IncompatibleReportSchema),
    ({"marker_overrides": {"B3": ""}}, IncompatibleReportSchema),
    ({"marker_overrides": {"B1": "=TODAY()"}}, IncompatibleReportSchema),
])
def test_structural_rejections(tmp_path, kwargs, error):
    with pytest.raises(error):
        _parse_bytes(tmp_path, build_ozon_products_workbook(**kwargs))


def test_extra_column_and_merged_cells_rejected(tmp_path):
    for mutation in ("extra", "merge"):
        book = load_workbook(BytesIO(build_ozon_products_workbook()))
        sheet = book.active
        if mutation == "extra": sheet.cell(row=5, column=33, value="extra")
        else: sheet.merge_cells("A7:B7")
        output = BytesIO(); book.save(output); book.close()
        with pytest.raises(IncompatibleReportSchema): _parse_bytes(tmp_path, output.getvalue())

def test_products_accepts_a1_stored_dimension_with_real_cells(tmp_path):
    assert len(_parse_bytes(tmp_path, build_ozon_products_workbook(dimension_ref="A1")).rows) == 1


def test_products_parser_never_uses_random_read_only_cell_access(tmp_path, monkeypatch):
    monkeypatch.setattr(ReadOnlyWorksheet, "cell", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("random access")))
    report = _parse_bytes(tmp_path, build_ozon_products_workbook(dimension_ref="A1"))
    assert len(report.rows) == 1


def test_products_realistic_report_parses_within_bound(tmp_path):
    rows = []
    for index in range(700):
        row = _default_row("Синтетическая категория")
        row[OZON_PRODUCTS_HEADERS[1]] = f"https://www.ozon.ru/product/{100000001 + index}/"
        rows.append(row)
    data = build_ozon_products_workbook(rows=rows, dimension_ref="A1")
    started = perf_counter()
    report = _parse_bytes(tmp_path, data)
    elapsed = perf_counter() - started
    assert (report.rows_seen, len(report.rows)) == (700, 700)
    assert report.rows[-1].ozon_product_id == "100000700"
    assert elapsed < 10


def test_recoverable_identity_and_duplicate_rows(tmp_path):
    valid = _default_row("Синтетическая категория")
    invalid = dict(valid); invalid[OZON_PRODUCTS_HEADERS[1]] = "bad"
    report = _parse_bytes(tmp_path, build_ozon_products_workbook(rows=[valid, valid, invalid]))
    assert len(report.rows) == 1 and report.duplicate_input_rows == 1 and report.warnings_count == 1
    assert report.rows_seen == 3
    assert report.row_errors == (RowError(9, "InvalidProductIdentity", "Некорректная ссылка на товар."),)


@pytest.mark.parametrize(
    ("header", "value"),
    [
        (OZON_PRODUCTS_HEADERS[7], "=1+1"),
        (OZON_PRODUCTS_HEADERS[7], "Нет данных"),
        (OZON_PRODUCTS_HEADERS[8], "-"),
        (OZON_PRODUCTS_HEADERS[7], "1,31"),
        (OZON_PRODUCTS_HEADERS[7], True),
        (OZON_PRODUCTS_HEADERS[31], "not-a-date"),
        (OZON_PRODUCTS_HEADERS[31], datetime(2026, 1, 1)),
        (OZON_PRODUCTS_HEADERS[6], 0),
    ],
)
def test_invalid_product_cells_are_recoverable(tmp_path, header, value):
    valid = _default_row("Синтетическая категория")
    invalid = dict(valid)
    invalid[OZON_PRODUCTS_HEADERS[1]] = "https://www.ozon.ru/product/100000002/"
    invalid[header] = value
    report = _parse_bytes(
        tmp_path, build_ozon_products_workbook(rows=[valid, invalid])
    )
    assert report.rows_seen == 2
    assert len(report.rows) == 1
    assert report.row_errors == (
        RowError(8, "InvalidMetricValue", "Некорректное значение показателя."),
    )


def test_exact_sentinels_zero_and_product_badges(tmp_path):
    headers = OZON_PRODUCTS_HEADERS
    rows = []
    for index, badge in enumerate((None, "", "text"), start=1):
        row = _default_row("Синтетическая категория")
        row[headers[1]] = f"https://www.ozon.ru/product/{100000000 + index}/"
        row[headers[6]] = badge
        row[headers[7]] = 0
        row[headers[8]] = "Нет данных"
        row[headers[12]] = "Нет данных"
        row[headers[14]] = "-"
        rows.append(row)
    report = _parse_bytes(tmp_path, build_ozon_products_workbook(rows=rows))
    assert [row.snapshot_values["product_badges"] for row in report.rows] == [
        None,
        None,
        "text",
    ]
    assert all(row.snapshot_values["ordered_amount_rub"] == 0 for row in report.rows)
    assert all(row.snapshot_values["turnover_change_pct"] is None for row in report.rows)
    assert all(row.snapshot_values["buyout_share_pct"] is None for row in report.rows)
    assert all(row.snapshot_values["out_of_stock_days"] is None for row in report.rows)


def test_conflicting_same_key_rows_are_fatal(tmp_path):
    first = _default_row("Синтетическая категория")
    changed = dict(first)
    changed[OZON_PRODUCTS_HEADERS[0]] = "Changed title"
    with pytest.raises(ConflictingObservationRows):
        _parse_bytes(tmp_path, build_ozon_products_workbook(rows=[first, changed]))
