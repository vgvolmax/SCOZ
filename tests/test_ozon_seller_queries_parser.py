from dataclasses import fields, FrozenInstanceError
from decimal import Decimal
import pytest
from openpyxl import load_workbook
from io import BytesIO
from backend.domain.product_query import *
from tests.xlsx_factory import build_ozon_seller_queries_workbook, OZON_SELLER_QUERIES_HEADERS

def test_contract_and_payload():
    assert tuple(f.name for f in fields(ProductQuerySnapshot))[11:] == PRODUCT_QUERY_PAYLOAD_FIELDS
    values = dict(searched_users=1, seen_users=1, position_state=ProductQueryPositionState.SOURCE_ZERO, average_position=None, search_to_card_conversion_pct=Decimal('2.480'), search_to_order_conversion_pct=Decimal('0'), ordered_units=0, ordered_revenue_rub=Decimal('0'))
    assert product_query_payload_sha256(values) == product_query_payload_sha256({**values, 'search_to_card_conversion_pct': Decimal('2.48')})
    with pytest.raises(TypeError): product_query_payload_sha256({**values, 'ordered_revenue_rub': 0.0})

def test_synthetic_shape():
    wb=load_workbook(BytesIO(build_ozon_seller_queries_workbook()))
    ws=wb.active
    assert tuple(ws.cell(6,c).value for c in range(1,12)) == OZON_SELLER_QUERIES_HEADERS
    assert all(ws.cell(5,c).value is None for c in range(1,12))
    wb.close()
